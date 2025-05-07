from datetime import datetime
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import TemplateView, View, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth import authenticate, login
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import OrderPaymentItem, UserProfile, User
from .forms import UserRegisterForm, OrderPaymentItemForm
from django.contrib import messages
from num2words import num2words
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import *
import openpyxl
from openpyxl.drawing.image import Image
import PIL
import io
import urllib3
from openpyxl.worksheet.cell_range import CellRange

# Create your views here.
class Index(TemplateView):
    template_name = 'order_payment/index.html'

class Dashboard(LoginRequiredMixin, View):
    def get(self, request):

        date_format = "%Y-%m-%d"
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        items = None

        if start_date and end_date:
            # Filter data within a date range
            items = OrderPaymentItem.objects.filter(date__range=(start_date, end_date))
            start_date = datetime.strptime(request.GET.get('start_date'), date_format).date()
            end_date = datetime.strptime(request.GET.get('end_date'), date_format).date()
        elif start_date:
            end_date = datetime.now().date()
            items = OrderPaymentItem.objects.filter(date__range=(start_date, end_date))
            start_date = datetime.strptime(request.GET.get('start_date'), date_format).date()
        else:
            start_date = datetime.now().date().replace(day=1)
            end_date = datetime.now().date()
            items = OrderPaymentItem.objects.filter(date__range=(start_date, end_date))

        new_order_payment = items.filter(
			status="NEW"
		)

        if new_order_payment.count() > 0:
            if new_order_payment.count() > 1:
                messages.info(request, f'There are {new_order_payment.count()} new items.')
            else:
                messages.info(request, f'There is {new_order_payment.count()} new item')

        new_order_payment_ids = OrderPaymentItem.objects.filter(
			status="NEW"
		).values_list('id', flat=True)

        return render(request, 'order_payment/dashboard.html/', {'items': items, 'new_order_payment_ids': new_order_payment_ids, 'start_date': start_date, 'end_date': end_date})

class SignUpView(View):
    def get(self, request):
        form = UserRegisterForm()
        return render(request, 'order_payment/signup.html', {'form': form})

    def post(self, request):
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            user = authenticate(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password1']
            )

            login(request, user)
            return redirect('index')
        
        return render(request, 'order_payment/signup.html', {'form': form})
    
def has_decimal(value):
    return isinstance(value, float) and value % 1 != 0

def num2words_pesos(amount):
        if has_decimal(amount):
            whole_number, centavos = str(amount).split('.')
            whole_number_words = num2words(int(whole_number))
            centavos_words = num2words(int(centavos))
            return f"{whole_number_words} pesos and {centavos_words} cents"
        else:
            whole_number_words = num2words(amount)
            return f"{whole_number_words} pesos"
    
class AddItem(LoginRequiredMixin, CreateView):
    model = OrderPaymentItem
    form_class = OrderPaymentItemForm
    template_name = 'order_payment/item_form.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        form.instance.user = self.request.user
        form.instance.created_by = self.request.user.username

        total = 0

        if form.instance.fee_type_1_amount is not None:
            total = total + form.instance.fee_type_1_amount
        
        if form.instance.fee_type_2_amount is not None:
            total = total + form.instance.fee_type_2_amount
        
        if form.instance.fee_type_3_amount is not None:
            total = total + form.instance.fee_type_3_amount

        if form.instance.dst is not None:
            total = total + form.instance.dst

        if form.instance.surcharge is not None:
            total = total + form.instance.surcharge

        form.instance.total_amount = total
        form.instance.amount_in_words = num2words_pesos(total)
        # form.instance.amount_in_words = num2words(total, to='currency', lang='en', separator=' and', currency='USD')

        return super().form_valid(form)

class EditItem(LoginRequiredMixin, UpdateView):
    model = OrderPaymentItem
    form_class = OrderPaymentItemForm
    template_name = 'order_payment/item_form.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        oldItem = OrderPaymentItem.objects.get(pk=form.instance.pk)
        form.instance.user = self.request.user
        form.instance.last_update_by = self.request.user.username

        if form.instance.status != oldItem.status and form.instance.status == 'APPROVED':
            form.instance.approver_username = self.request.user.username
            form.instance.signature_url = self.request.user.userprofile.signature.url
            form.instance.approver_name = self.request.user.first_name + " " + self.request.user.last_name
        
        if form.instance.status == 'NEW':
            form.instance.approver_username = None
            form.instance.signature_url = None
            form.instance.approver_name = None

        total = 0

        if form.instance.fee_type_1_amount is not None:
            total = total + form.instance.fee_type_1_amount
        
        if form.instance.fee_type_2_amount is not None:
            total = total + form.instance.fee_type_2_amount
        
        if form.instance.fee_type_3_amount is not None:
            total = total + form.instance.fee_type_3_amount

        if form.instance.dst is not None:
            total = total + form.instance.dst

        if form.instance.surcharge is not None:
            total = total + form.instance.surcharge

        form.instance.total_amount = total
        form.instance.amount_in_words = num2words_pesos(total)

        return super().form_valid(form)

class DeleteItem(LoginRequiredMixin, DeleteView):
	model = OrderPaymentItem
	template_name = 'order_payment/delete_item.html'
	success_url = reverse_lazy('dashboard')
	context_object_name = 'item'

class ViewItem(LoginRequiredMixin, DetailView):
    model = OrderPaymentItem
    template_name = 'order_payment/view_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["order_payment"] = self.get_object
        return context

def export_item_excel(request, pk=None):

    item = OrderPaymentItem.objects.get(pk=pk)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + item.serial_number +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = item.serial_number

    col_pos = 2
    row_pos = 2

    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Entity Name:"
    cell.font  = Font(bold=True)
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.entity_name
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font  = Font(bold=True)

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Fund Cluster:"
    cell.font  = Font(bold=True)
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fund_cluster
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font  = Font(bold=True)

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Serial No:"
    cell.font  = Font(bold=True)
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.serial_number
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font  = Font(bold=True)

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Date:"
    cell.font  = Font(bold=True)
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.date
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font  = Font(bold=True)

    row_pos += 2
    worksheet.merge_cells('B7:C7')
    title_cell = worksheet.cell(row=row_pos, column=col_pos)
    title_cell.value = "ORDER OF PAYMENT"
    title_cell.font  = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    row_pos += 2
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "The Collecting Officer"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.collecting_officer
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 2
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Please issue Official Receipt in favor of:"

    row_pos += 2
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Payor:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.payor
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Address:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.address
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Purpose/Type of Fee (1):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_1
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Amount of Fee (1):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_1_amount
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Purpose/Type of Fee (2):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_2
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Amount of Fee (2):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_2_amount
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Purpose/Type of Fee (3):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_3
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Amount of Fee (13):"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.fee_type_3_amount
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "DST:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.dst
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Surcharge:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.surcharge
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Total Amount:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.total_amount
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Amount in Words:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.amount_in_words
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Bill No.:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.bill_no
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Bill Date:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.bill_date
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 2
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Please deposit the collections under Bank Account/s:"

    row_pos += 2
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Name of Bank:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.bank_name
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Account Number:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.account_number
    cell.alignment = Alignment(horizontal="left", vertical="center")

    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Amount:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.deposit_amount
    cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in worksheet.columns:
        colLen = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
             if len(str(cell.value)) > colLen:
                 colLen = len(str(cell.value))
        set_col_width = colLen + 3
        # setting the column width
        worksheet.column_dimensions[column].width = set_col_width

    worksheet.column_dimensions["A"].width = 3

    if item.signature_url is not None:
        http = urllib3.PoolManager()
        r = http.request('GET', item.signature_url)
        image_file = io.BytesIO(r.data)
        img = Image(image_file)
        img.height = 40
        img.width = 120
        worksheet.add_image(img, 'C34')

    row_pos += 4
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Approver:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = item.approver_name
    cell.alignment = Alignment(horizontal="left", vertical="center")

    user = User.objects.get(username=item.approver_username)
    row_pos += 1
    cell = worksheet.cell(row=row_pos, column=col_pos)
    cell.value = "Role:"
    cell = worksheet.cell(row=row_pos, column=col_pos+1)
    cell.value = user.userprofile.role
    cell.alignment = Alignment(horizontal="left", vertical="center")

    range = CellRange("B2:C37")
    for row, col in range.cells:
        top = Side(style="thin") if (row, col) in range.top else None
        left = Side(style="thin") if (row, col) in range.left else None
        right = Side(style="thin") if (row, col) in range.right else None
        bottom = Side(style="thin") if (row, col) in range.bottom else None
        worksheet.cell(row, col).border = Border(left, right, top, bottom, outline=True)

    workbook.save(response)
    return response